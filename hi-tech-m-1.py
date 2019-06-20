# -*- coding: UTF-8 -*-
import os
import os.path
import logging
import logging.config
import sys
import configparser
import time
import shutil
import openpyxl                       # Для .xlsx
#import xlrd                          # для .xls
from   price_tools import getCellXlsx, getCell, nameToId, currencyTypeX, sheetByName
import csv



def getXlsString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]-1
        if item in ('закупка','продажа','цена1') :
            if getCell(row=i, col=j, isDigit='N', sheet=sh).find('По запросу') >=0 :
                impValues[item] = '0.1'
            else :
                impValues[item] = getCell(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            impValues[item] = currencyType(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCell(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def getXlsxString(sh, i, in_columns_j):
    impValues = {}
    for item in in_columns_j.keys() :
        j = in_columns_j[item]
        if item in ('закупка','продажа','цена2','цена1') :
            sss = getCellXlsx(row=i, col=j, isDigit='N', sheet=sh)
            if (sss.find('запросу') >=0 or sss.find('Звоните') >= 0):
                impValues[item] = '0.1'
            else :
                impValues[item] = getCellXlsx(row=i, col=j, isDigit='Y', sheet=sh)
            #print(sh, i, sh.cell( row=i, column=j).value, sh.cell(row=i, column=j).number_format, currencyType(sh, i, j))
        elif item == 'валюта_по_формату':
            #impValues[item] = 'RUR'
            impValues[item] = currencyTypeX(row=i, col=j, sheet=sh)
        else:
            impValues[item] = getCellXlsx(row=i, col=j, isDigit='N', sheet=sh)
    return impValues



def convert_excel2csv(cfg):
    priceFName= cfg.get('basic','filename_in')
    sheetName = cfg.get('basic','sheetname')


    log.debug('Reading file ' + priceFName )
    book, sheet = sheetByName(fileName = priceFName, sheetName = sheetName)
    if not sheet:
        log.error("Нет листа "+sheetName+" в файле "+ priceFName)
        return False
    log.debug("Sheet   "+sheetName)
    out_cols = cfg.options("cols_out")
    in_cols  = cfg.options("cols_in")
    out_template = {}
    for vName in out_cols :
         out_template[vName] = cfg.get("cols_out", vName)
    in_cols_j = {}
    for vName in in_cols :
         in_cols_j[vName] = cfg.getint("cols_in",  vName)
    #brands,   discount     = config_read(cfgFName, 'discount')
    #for k in discount.keys():
    #    discount[k] = (100 - int(discount[k]))/100
    #print(discount)

    outFileUSD = False
    outFileEUR = False
    outFileRUR = False
    if cfg.has_option('basic','filename_out_RUR'):
        csvFfileNameRUR = cfg.get('basic', 'filename_out_RUR')
        outFileRUR = open(csvFfileNameRUR, 'w', newline='')
        csvWriterRUR = csv.DictWriter(outFileRUR, fieldnames=cfg.options('cols_out'))
        csvWriterRUR.writeheader()
    if cfg.has_option('basic', 'filename_out_USD'):
        csvFfileNameUSD = cfg.get('basic', 'filename_out_USD')
        outFileUSD = open(csvFfileNameUSD, 'w', newline='')
        csvWriterUSD = csv.DictWriter(outFileUSD, fieldnames=cfg.options('cols_out'))
        csvWriterUSD.writeheader()
    if cfg.has_option('basic', 'filename_out_EUR'):
        csvFfileNameEUR = cfg.get('basic', 'filename_out_EUR')
        outFileEUR = open(csvFfileNameEUR, 'w', newline='')
        csvWriterEUR = csv.DictWriter(outFileEUR, fieldnames=cfg.options('cols_out'))
        csvWriterEUR.writeheader()


    '''                                     # Блок проверки свойств для распознавания групп      XLSX
    for i in range(2, 15):
        i_last = i
        ccc = sheet.cell( row=i, column=in_cols_j['подгруппа'] )
        print(i, sheet.cell(row=i, column=in_cols_j['цена1']).value, ccc.value)
        print(ccc.font.name, ccc.font.sz, ccc.font.b, ccc.font.i, '------', 'ccc.font.color.rgb', ccc.fill.bgColor.rgb, 'ccc.fill.fgColor.rgb')
        print('------')
    return
    '''
    '''                                     # Блок проверки свойств для распознавания групп      XLS                                  
    for i in range(19, 25):                                                         
        xfx = sheet.cell_xf_index(i, 1)
        xf  = book.xf_list[xfx]
        bgci  = xf.background.pattern_colour_index
        fonti = xf.font_index
        ccc = sheet.cell(i, 1)
        if ccc.value == None :
            print (i, colSGrp, 'Пусто!!!')
            continue
                                         # Атрибуты шрифта для настройки конфига
        font = book.font_list[fonti]
        print( '---------------------- Строка', i, '-----------------------', sheet.cell(i, 1).value)
        print( 'background_colour_index=',bgci)
        print( 'fonti=', fonti, '           xf.alignment.indent_level=', xf.alignment.indent_level)
        print( 'bold=', font.bold)
        print( 'weight=', font.weight)
        print( 'height=', font.height)
        print( 'italic=', font.italic)
        print( 'colour_index=', font.colour_index )
        print( 'name=', font.name)
    return
    '''

    recOut  ={}
    subgrp = ''
    for i in range(1, sheet.max_row +1) :                               # xlsx
#   for i in range(1, sheet.nrows) :                                     # xls
        i_last = i
        try:
#            xfx = sheet.cell_xf_index(i, 1)                              # xls
#            xf  = book.xf_list[xfx]                                      # xls
#            bgci  = xf.background.pattern_colour_index                   # xls
            impValues = getXlsxString(sheet, i, in_cols_j)                # xlsx
            #impValues = getXlsString(sheet, i, in_cols_j)                # xls
            #print( impValues )
            ccc1 = sheet.cell(row=i, column=in_cols_j['цена1']).value

            if sheetName == 'Antall':
                if (sheet.cell(row=i, column=in_cols_j['подгруппа']).font.b is True and
                    sheet.cell(row=i, column=in_cols_j['цена1']).value is None):          # подгруппа
                    subgrp = impValues['подгруппа']
                    continue
                elif (impValues['код_'] == '' or
                    impValues['код_'] == 'Модель' or
                    impValues['цена1'] == '0'):                                           # лишняя строка
                    continue
                impValues['подгруппа'] = subgrp

            elif sheetName == 'LG':
                if (sheet.cell(row=i, column=in_cols_j['группа_']).font.b is True and
                    sheet.cell(row=i, column=in_cols_j['цена1']).value is None):          # группа
                    grp = impValues['группа_']
                    subgrp = ''
                    continue
                elif (sheet.cell(row=i, column=in_cols_j['подгруппа']).font.b is True and
                    sheet.cell(row=i, column=in_cols_j['цена1']).value is not None):      # подгруппа
                    subgrp = impValues['подгруппа']
                elif (impValues['код_'] == '' or
                    impValues['код_'] == 'Модель' or
                    impValues['цена1'] == '0'):                                           # лишняя строка
                    continue
                impValues['группа_'] = grp
                impValues['подгруппа'] = subgrp

            elif sheetName == 'ArthurHolm':
                if (sheet.cell(row=i, column=in_cols_j['подгруппа']).font.b is True and
                    sheet.cell(row=i, column=in_cols_j['цена1']).value is None):          # подгруппа
                    subgrp = impValues['подгруппа']
                    continue
                elif (impValues['код_'] == '' or
                    impValues['код_'] == 'Модель' or
                    impValues['цена1'] == '0'):                                           # лишняя строка
                    continue
                impValues['подгруппа'] = subgrp
                if '\n' in impValues['код_']:
                    p = impValues['код_'].find('\n')
                    impValues['код_'] = impValues['код_'][:p]
            else:
                log.error('нераспознан sheetName "%s"', sheetName)      # далее общая для всех обработка

            for outColName in out_template.keys() :
                shablon = out_template[outColName]
                for key in impValues.keys():
                    if shablon.find(key) >= 0:
                        shablon = shablon.replace(key, impValues[key])
                if (outColName == 'закупка') and ('*' in shablon) :
                    p = shablon.find("*")
                    vvv1 = float(shablon[:p])
                    vvv2 = float(shablon[p+1:])
                    shablon = str(round(vvv1 * vvv2, 2))
                recOut[outColName] = shablon.strip()

            recOut['код'] = nameToId(recOut['код'])
            if recOut['валюта'] != 'USD' and recOut['продажа'] == '0.1':
                recOut['валюта'] = 'USD'
                recOut['закупка'] = '0.1'
            if recOut['валюта'] == 'RUR':
                csvWriterRUR.writerow(recOut)
            elif recOut['валюта'] == 'USD':
                csvWriterUSD.writerow(recOut)
            elif recOut['валюта'] == 'EUR':
                csvWriterEUR.writerow(recOut)
            else:
                log.error('нераспознана валюта "%s" для товара "%s"', recOut['валюта'], recOut['код производителя'])

        except Exception as e:
            print(e)
            if str(e) == "'NoneType' object has no attribute 'rgb'":
                pass
            else:
                log.debug('Exception: <' + str(e) + '> при обработке строки ' + str(i) +'.' )

    log.info('Обработано ' +str(i_last)+ ' строк.')
    if outFileRUR:
        outFileRUR.close()
    if outFileUSD:
        outFileUSD.close()
    if outFileEUR:
        outFileEUR.close()



def download( cfg ):
    from selenium import webdriver
    from selenium.webdriver.common.keys import Keys
    from selenium.webdriver.remote.remote_connection import LOGGER
    LOGGER.setLevel(logging.WARNING)
     
    retCode     = False
    filename_new_1= cfg.get('basic','filename_new_1')
    filename_old_1= cfg.get('basic','filename_old_1')
    filename_new_2= cfg.get('basic','filename_new_2')
    filename_old_2= cfg.get('basic','filename_old_2')
    #login       = cfg.get('download','login'    )
    #password    = cfg.get('download','password' )
    #url_lk      = cfg.get('download','url_lk'   )
    url_file_1 = cfg.get('download','url_file_1')
    url_file_2 = cfg.get('download','url_file_2')

    download_path= os.path.join(os.getcwd(), 'tmp')
    if not os.path.exists(download_path):
        os.mkdir(download_path)

    for fName in os.listdir(download_path) :
        os.remove( os.path.join(download_path, fName))
    dir_befo_download = set(os.listdir(download_path))
        
    if os.path.exists('geckodriver.log') : os.remove('geckodriver.log')
    try:
        ffprofile = webdriver.FirefoxProfile()
        ffprofile.set_preference("browser.download.dir", download_path)
        ffprofile.set_preference("browser.download.folderList",2);
        ffprofile.set_preference("browser.helperApps.neverAsk.saveToDisk", 
                ",application/octet-stream" + 
                ",application/vnd.ms-excel" + 
                ",application/vnd.msexcel" + 
                ",application/x-excel" + 
                ",application/x-msexcel" + 
                ",application/zip" + 
                ",application/xls" + 
                ",application/vnd.ms-excel" +
                ",application/vnd.ms-excel.addin.macroenabled.12" +
                ",application/vnd.ms-excel.sheet.macroenabled.12" +
                ",application/vnd.ms-excel.template.macroenabled.12" +
                ",application/vnd.ms-excelsheet.binary.macroenabled.12" +
                ",application/vnd.ms-fontobject" +
                ",application/vnd.ms-htmlhelp" +
                ",application/vnd.ms-ims" +
                ",application/vnd.ms-lrm" +
                ",application/vnd.ms-officetheme" +
                ",application/vnd.ms-pki.seccat" +
                ",application/vnd.ms-pki.stl" +
                ",application/vnd.ms-word.document.macroenabled.12" +
                ",application/vnd.ms-word.template.macroenabed.12" +
                ",application/vnd.ms-works" +
                ",application/vnd.ms-wpl" +
                ",application/vnd.ms-xpsdocument" +
                ",application/vnd.openofficeorg.extension" +
                ",application/vnd.openxmformats-officedocument.wordprocessingml.document" +
                ",application/vnd.openxmlformats-officedocument.presentationml.presentation" +
                ",application/vnd.openxmlformats-officedocument.presentationml.slide" +
                ",application/vnd.openxmlformats-officedocument.presentationml.slideshw" +
                ",application/vnd.openxmlformats-officedocument.presentationml.template" +
                ",application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" +
                ",application/vnd.openxmlformats-officedocument.spreadsheetml.template" +
                ",application/vnd.openxmlformats-officedocument.wordprocessingml.template" +
                ",application/x-ms-application" +
                ",application/x-ms-wmd" +
                ",application/x-ms-wmz" +
                ",application/x-ms-xbap" +
                ",application/x-msaccess" +
                ",application/x-msbinder" +
                ",application/x-mscardfile" +
                ",application/x-msclip" +
                ",application/x-msdownload" +
                ",application/x-msmediaview" +
                ",application/x-msmetafile" +
                ",application/x-mspublisher" +
                ",application/x-msschedule" +
                ",application/x-msterminal" +
                ",application/x-mswrite" +
                ",application/xml" +
                ",application/xml-dtd" +
                ",application/xop+xml" +
                ",application/xslt+xml" +
                ",application/xspf+xml" +
                ",application/xv+xml" +
                ",application/excel")
        if os.name == 'posix':
            #driver = webdriver.Firefox(ffprofile, executable_path=r'/usr/local/Cellar/geckodriver/0.19.1/bin/geckodriver')
            driver = webdriver.Firefox(ffprofile, executable_path=r'/usr/local/bin/geckodriver')
        elif os.name == 'nt':
            driver = webdriver.Firefox(ffprofile)
        driver.implicitly_wait(10)
        driver.set_page_load_timeout(10)

        try:
            driver.get(url_file_1)
        except Exception as e:
            log.debug('Exception: <' + str(e) + '>')
        dir_afte_download = set(os.listdir(download_path))
        new_files = list(dir_afte_download.difference(dir_befo_download))
        print(new_files)
        if len(new_files) < 1:
            log.error('Не удалось скачать файл прайса ' + filename_new_1)
            retCode = False
        elif len(new_files) > 1:
            log.error('Скачалось несколько файлов. Надо разбираться ...')
            retCode = False
        else:
            new_file_1 = new_files[0]
            new_ext_1 = os.path.splitext(new_file_1)[-1].lower()
            DnewFile1 = os.path.join(download_path, new_file_1)
            new_file_date = os.path.getmtime(DnewFile1)
            log.info('Скачанный файл ' + new_file_1 + ' имеет дату ' + time.strftime("%Y-%m-%d %H:%M:%S",
                                                                                   time.localtime(new_file_date)))
        '''
        dir_befo_download = set(os.listdir(download_path))
        try:
            driver.get(url_file_2)
        except Exception as e:
            log.debug('Exception: <' + str(e) + '>')
        '''
    except Exception as e:
        log.debug('Exception: <' + str(e) + '>')
    driver.quit()
    '''
    dir_afte_download = set(os.listdir(download_path))
    new_files = list(dir_afte_download.difference(dir_befo_download))
    if len(new_files) < 1:
        log.error('Не удалось скачать файл прайса ' + filename_new_2)
        retCode = False
    elif len(new_files) > 1:
        log.error('Скачалось несколько файлов. Надо разбираться ...')
        retCode = False
    else:
        new_file = new_files[0]                                                     # загруженo ровно 1 файл.
        new_ext_2  = os.path.splitext(new_file)[-1].lower()
        DnewFile2 = os.path.join( download_path,new_file)
        new_file_date = os.path.getmtime(DnewFile2)
        log.info('Скачанный файл ' +new_file + ' имеет дату ' + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(new_file_date) ) )

    if os.path.exists(filename_new_2) and os.path.exists(filename_old_2):
        os.remove(filename_old_2)
        os.rename(filename_new_2, filename_old_2)
    if os.path.exists(filename_new_2):
        os.rename(filename_new_2, filename_old_2)
    shutil.copy2(DnewFile2, filename_new_2)
    retCode = True
    '''
    if new_ext_1 == '.zip':
        log.debug('Zip-архив. Разархивируем.')
        work_dir = os.getcwd()
        os.chdir(os.path.join(download_path))
        dir_befo_download = set(os.listdir(os.getcwd()))
        os.system('unzip -oj ' + new_file_1)
        os.remove(new_file_1)
        dir_afte_download = set(os.listdir(os.getcwd()))
        new_files = list(dir_afte_download.difference(dir_befo_download))
        os.chdir(work_dir)
        if len(new_files) == 1:
            new_file = new_files[0]  # разархивирован ровно один файл.
            new_ext = os.path.splitext(new_file)[-1]
            DnewFile = os.path.join(download_path, new_file)
            new_file_date = os.path.getmtime(DnewFile)
            log.debug('Файл из архива ' + DnewFile + ' имеет дату ' + time.strftime("%Y-%m-%d %H:%M:%S",
                                                                                time.localtime(new_file_date)))
            if os.path.exists(filename_new_1) and os.path.exists(filename_old_1):
                os.remove(filename_old_1)
                os.rename(filename_new_1, filename_old_1)
            if os.path.exists(filename_new_1):
                os.rename(filename_new_1, filename_old_1)
            shutil.copy2(DnewFile, filename_new_1)
            retCode = True

        elif len(new_files) > 1:
            log.debug('В архиве не единственный файл. Надо разбираться.')
            retCode = False
        else:
            log.debug('Нет новых файлов после разархивации. Загляни в папку юниттеста поставщика.')
            retCode = False

    return retCode




def config_read( cfgFName ):
    cfg = configparser.ConfigParser(inline_comment_prefixes=('#'))
    if  os.path.exists('private.cfg'):     
        cfg.read('private.cfg', encoding='utf-8')
    if  os.path.exists('getting.cfg'):
        cfg.read('getting.cfg', encoding='utf-8')
    if  os.path.exists(cfgFName):
        cfg.read( cfgFName, encoding='utf-8')
    else: 
        log.debug('Нет файла конфигурации '+cfgFName)
    return cfg



def is_file_fresh(fileName, qty_days):
    qty_seconds = qty_days *24*60*60 
    if os.path.exists( fileName):
        price_datetime = os.path.getmtime(fileName)
    else:
        log.error('Не найден файл  '+ fileName)
        return False

    file_age = round((time.time() - price_datetime) / 24 / 60 / 60)
    if file_age > qty_days :
        log.error('Файл "' + fileName + '" устарел! Допустимый период ' + str(qty_days)+' дней, а ему ' + str(file_age))
        return False
    else:
        return True



def make_loger():
    global log
    logging.config.fileConfig('logging.cfg')
    log = logging.getLogger('logFile')


def main(dealerName):
    """ Обработка прайсов выполняется согласно файлов конфигурации.
    Для этого в текущей папке должны быть файлы конфигурации, описывающие
    свойства файла и правила обработки. По одному конфигу на каждый
    прайс или раздел прайса со своими правилами обработки
    """
    make_loger()
    log.info('          ' + dealerName)

    rc_download = False
    '''
    if os.path.exists('getting.cfg'):
        cfg = config_read('getting.cfg')
        filename_new_1 = cfg.get('basic','filename_new_1')
        filename_new_2 = cfg.get('basic','filename_new_2')
        if cfg.has_section('download'):
            rc_download = download(cfg)
        if not(rc_download==True or is_file_fresh( filename_new_1, int(cfg.get('basic','срок годности')))):
            return False
    '''
    for cfgFName in os.listdir("."):
        if cfgFName.startswith("cfg") and cfgFName.endswith(".cfg"):
            log.info('----------------------- Processing '+cfgFName )
            cfg = config_read(cfgFName)
            filename_in = cfg.get('basic','filename_in')
            if rc_download==True or is_file_fresh( filename_in, int(cfg.get('basic','срок годности'))):
                convert_excel2csv(cfg)



if __name__ == '__main__':
    myName = os.path.basename(os.path.splitext(sys.argv[0])[0])
    mydir    = os.path.dirname (sys.argv[0])
    print(mydir, myName)
    main( myName)
